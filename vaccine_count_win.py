"""
Created on Thu Sep  7 16:57:43 2021
코로나19 접종예약명단(모더나포함)을 csv파일로 받아서 xlsx로 변환하고 정리해주는 스크립트 입니다.
본 스크립트는 누구나 사용가능합니다.
최종 결과물이 마음에 안드시면 마음껏 고쳐서 사용해 주세요
본 스크립트는 GPL 을 따릅니다.
@author: wonsungchulmd@gmail.com
"""
import openpyxl
import pandas as pd
import datetime
import os
import tkinter
import sys
from tkinter import Tk
from tkinter import filedialog
# cell border style, font bold, font normal, center 정렬, cell 음영 정의
thin = openpyxl.styles.Side(border_style="thin", color="000000")
thick = openpyxl.styles.Side(border_style="thick", color="000000")
double = openpyxl.styles.Side(border_style="double", color="000000")
border = openpyxl.styles.Border(left=thin, right=thin, top=thin, bottom=thin)
border_thick = openpyxl.styles.Border(left=thick, right=thick, top=thick, bottom=thick)
border_btm_thick = openpyxl.styles.Border(left=thin, right=thin, top=thin, bottom=thick)
border_btm_double = openpyxl.styles.Border(left=thin, right=thin, top=thin, bottom=double)
align_center = openpyxl.styles.Alignment(horizontal='center')
font_bold = openpyxl.styles.Font(bold=True)
font_nl = openpyxl.styles.Font(bold=False)
c_fill2 = openpyxl.styles.PatternFill(start_color='fff999', end_color='fff999', fill_type='solid')
c_fill = openpyxl.styles.PatternFill(start_color='ff9999', end_color='ff9999', fill_type='solid')
try:
    ## 파일이름, 경로 변수 저장
    user_def = os.path.join(os.path.expanduser('~'),'Desktop') + '\\'
    root = Tk()
    root.filename = filedialog.askopenfilename(initialdir = user_def,title="Choose your file", filetypes = (("csv files","*.csv"),("all files","*.*")))
    #다운로드받은 파일이름이 1.csv가 아닌경우 변경해주세요
    #저장하고 싶은 엑셀파일 이름이 있으면 변경해주세요. 꼭 확장자까지 변경해야 합니다.
    ########################## 파일명 변경 ##########################################
    csv_file_name = os.path.basename(root.filename)
    #xlsx_file_name_1 = '1.xlsx'
    xlsx_file_name_2 = '최종본.xlsx'
    ###############################################################################
    ################# 토요일 및 평일 upper limit 환자 수 정의 #####################
    sat_u_cnt = 55
    weekday_u_cnt = 85
    ###############################################################################
    eng_weekday = ['Mon','Tue','Wed','Thu','Fri','Sat','Sun']
    kor_weekday = ['월요일','화요일','수요일','목요일','금요일','토요일','일요일']
    ## 파일이름, 경로 변수 저장
    #user_def = os.path.join(os.path.expanduser('~'),'Desktop')
    path= user_def + csv_file_name
    #path_xlsx_1 = user_def + xlsx_file_name_1
    path_xlsx_2 = user_def + xlsx_file_name_2
    col_alphabet = ['A','B','C','D','E','F','G']
    width_fix = (13,8,14,14,9,14.5,19.5)
    col_alphabet2 = ['A','B','C','D','E','F','G','H']
    width_fix2 = (13,8,8,8,11,11,14.5,8)
    # 해당이름의 csv파일을 읽어옴(1.csv)
    try:
        r_csv = pd.read_csv(path, encoding='euc-kr')
    except:
        print("Unexpected error:", sys.exc_info()[0])
    # 저장할 xlsx파일의 이름을 정함(1.xlsx)
    save_xlsx = pd.ExcelWriter(path_xlsx_2)
    # xlsx 파일로 변환
    r_csv.to_excel(save_xlsx, index = False) 
    #xlsx 파일로 저장
    save_xlsx.save() 
    # 바탕화면의 1.xlsx에서 active sheet을 불러와서 워크시트 이름을 sh_name에 넣음
    wb = openpyxl.load_workbook(filename = path_xlsx_2)
    ws = wb.active
    sh_name = ws.title
    #csv 읽을때 접종자 이름 오류부분(같은 이름이 있을때 판다스에서 컬럼name이 겹쳐서 변경된 부분) 정정코드
    w_text=ws.cell(1,6).value
    r_text=w_text.replace(".1","")
    ws.cell(1,6, value = r_text)
    #active worksheet에서 필요없는 컬럼 제거
    del_col_no = (12,11,10,9,4,1)
    for i in del_col_no:
        ws.delete_cols(i)
    #요일 넣을 컬럼 삽입
    ws.insert_cols(2)
    #pandas에서 첫줄을 컬럼명으로 인지해서 굵게 되어있는 부분 정상화
    for i in range(1,8):
        ws.cell(1,i).font = font_nl
    #전체 데이터 시트(ws,Sheet1)에서 제목을 첫째행 맨 위에 삽입 후 넓이 조절, 정렬 가운데, 폰트 굵게 + thin border 적용
    ws.insert_rows(1)
    subject1 = ['날자','요일','시간','전화번호','이름','주민번호','접종종류']
    col_no = range(1,8)
    for i, j in zip(col_no, subject1):
        ws.cell(1,i, value=j)
        ws.cell(1,i).border = border_btm_thick
        ws.cell(1,i).alignment = align_center
        ws.cell(1,i).font = font_bold
    for i, j in zip(col_alphabet, width_fix):
        ws.column_dimensions[i].width = j
    #오늘날자 명단 시트 생성(ws1) 및 첫째행에 제목삽입 후 넓이 조절
    wb.create_sheet('오늘환자명단',0)
    ws1 = wb['오늘환자명단']
    subject1 = ['날자','요일','시간','전화번호','이름','주민번호','접종종류']
    col_no = range(1,8)
    for i, j in zip(col_no, subject1):
        ws1.cell(1,i, value=j)
        ws1.cell(1,i).border = border_btm_thick
        ws1.cell(1,i).alignment = align_center
        ws1.cell(1,i).font = font_bold
        ws1.cell(1,i).fill=c_fill2 
    for i, j in zip(col_alphabet, width_fix):
        ws1.column_dimensions[i].width = j
    #백신종류별환자수시트 생성(ws2)후 첫째행에 제목삽입, 제목행 굵게, 밑줄 굵게
    wb.create_sheet('백신종류별환자수',1)
    ws2 = wb['백신종류별환자수']
    subject1 = ['날자','요일','총인원','화이자','Pf_vial(6)','pf_vial(7)','아스트라제네카','모더나']
    col_no = range(1,9)
    for i, j in zip(col_no, subject1):
        ws2.cell(1,i, value=j)
        ws2.cell(1,i).font = font_bold
        ws2.cell(1,i).border = border_btm_thick
        ws2.cell(1,i).alignment = align_center
    for i, j in zip(col_alphabet2, width_fix2):
        ws2.column_dimensions[i].width = j
    #전체데이터시트(ws,Sheet1)에서 날자 포맷변경, 요일획득후 빈 공간에 날자 채우기
    for i in range(2,ws.max_row+1):
        if (ws.cell(i,1).value != None) and (ws.cell(i,1).value == ws.cell(i+1,1).value):
            r_date = ws.cell(i,1).value #str 타입 날자를 r_date에 넣음
            dateFormatter = "%Y-%m-%d" #날자 포맷 형식
            w_date=datetime.datetime.strptime(r_date, dateFormatter) #str 타입 날자를 date/time타입 날자로 변경(변수 w_date)
            new_date=w_date.strftime("%Y-%m-%d")#date/time 형식 날자를 YYYY-mm-dd형식의 str 타입으로 변경
            new_weekday=w_date.strftime("%a")#w_date에서 요일 'Wed'타입으로 불러 new_weekday에 넣음(str 형식)
            ws.cell(i,1, value=new_date)
            ws.cell(i,2, value=new_weekday)#두번째 cell에 요일을 넣음
            for j in range(1,8):
                ws.cell(i,j).border = border #8번째 셀까지 전체 thin border
        elif  (ws.cell(i,1).value != None) and (ws.cell(i+1,1).value == None):
            r_date = ws.cell(i,1).value #str 타입 날자를 r_date에 넣음
            dateFormatter = "%Y-%m-%d" #날자 포맷 형식
            w_date=datetime.datetime.strptime(r_date, dateFormatter) #str 타입 날자를 date/time타입 날자로 변경(변수 w_date)
            new_date=w_date.strftime("%Y-%m-%d")#date/time 형식 날자를 YYYY-mm-dd형식의 str 타입으로 변경
            new_weekday=w_date.strftime("%a")#w_date에서 요일 'Wed'타입으로 불러 new_weekday에 넣음(str 형식)
            ws.cell(i,1, value=new_date)
            ws.cell(i,2, value=new_weekday)#두번째 cell에 요일을 넣음
            ws2.append([new_date,new_weekday])
            for j in range(1,8):
                ws.cell(i,j).border = border #8번째 셀까지 전체 thin border
        else: #날자 cell이 비어있는 경우에는 위에서 저장한 변수를 빈 cell에 넣기
            ws.cell(i,1, value=new_date)
            ws.cell(i,2, value=new_weekday)
            for j in range(1,8):
                ws.cell(i,j).border = border #변수 채운 후 8번째 cell까지 thin border 적용
    #엑셀의 COUNTIF함수 동작
    k2=2
    q=2
    for k in range(2, ws.max_row+1):
        if ws.cell(k,1).value != ws.cell(k+1,1).value:
            n_row = k
            l = 0
            m = 0
            n = 0
            for o in range(k2,n_row+1):
                if (ws.cell(o,7).value == '1차 (화이자)') or (ws.cell(o,7).value == '2차 (화이자)'):
                    l += 1
                elif(ws.cell(o,7).value == '1차 (아스트라제네카)') or (ws.cell(o,7).value == '2차 (아스트라제네카)'):
                    m += 1
                elif(ws.cell(o,7).value == '1차 (모더나)') or (ws.cell(o,7).value == '2차 (모더나)'):
                    n += 1
                k3 = k-k2+1
                y=(l//6)+1 # 화이자바이알계산(6개씩 뽑을때)
                z=(l//7)+1 # 화이자바이알계산(7개씩 뽑을때)
            for r,vac_value in zip(range(3,9),(k3,l,y,z,m,n)):
                ws2.cell(q,r, value=vac_value)
            k2=n_row+1
            q += 1
    #시간셀에 값 확인해서 없는 경우 넣어주기
    for i in range(2,ws.max_row+1):
        if (ws.cell(i,3).value != None) and (ws.cell(i,3).value == ws.cell(i+1,3).value):
            new_time = ws.cell(i,3).value
            ws.cell(i,3, value=new_time)
            for j in range(1,8):
                ws.cell(i,j).alignment = align_center
        elif (ws.cell(i,3).value != None) and (ws.cell(i+1,3).value == None):
            new_time = ws.cell(i,3).value
            ws.cell(i,3, value=new_time)
            for j in range(1,8):
                ws.cell(i,j).alignment = align_center
        else:
            ws.cell(i,3, value=new_time)
            for j in range(1,8):
                ws.cell(i,j).alignment = align_center
                
    if ws.cell(ws.max_row,2).value == None:
        ws.cell(ws.max_row,3).value =''
    #전체명단시트(Sheet1) 날자 바뀌면 굵은 밑줄 넣어주기
    for i in range(2,ws.max_row+1):
        if ws.cell(i,1).value != ws.cell(i+1,1).value:
            for j in range(1,8):
                ws.cell(i,j).border = border_btm_thick
    #영어요일 => 한국어 변경 코딩
    ws_range = range(2,ws.max_row+1)
    ws1_range = range(2,ws1.max_row+1)
    ws2_range = range(2,ws2.max_row+1)
    for i in ws_range:
        for j,k in zip(eng_weekday,kor_weekday):
            if ws.cell(i,2).value == j:
                ws.cell(i,2, value = k)
    for i in ws1_range:
        for j,k in zip(eng_weekday,kor_weekday):
            if ws1.cell(i,2).value == j:
                ws1.cell(i,2, value = k)
    for i in ws2_range:
        for j,k in zip(eng_weekday,kor_weekday):
            if ws2.cell(i,2).value == j:
                ws2.cell(i,2, value = k)
    #백신종류별 환자수 시트에서 중간맞춤, thin_border 적용        
    for i in range(2,ws2.max_row+1):
        for j in range(1,9):
            ws2.cell(i,j).alignment = align_center
            ws2.cell(i,j).border = border
    #백신종류별 환자수 시트의 요일셀 중 토요일과 월요일이 포함된 행에 color_fill
    for i in range(2,ws2.max_row+1):
        if ws2.cell(i,2).value =='토요일':
            for j in range(1,9):
                ws2.cell(i,j).fill=c_fill
                ws2.cell(i,j).font=font_bold
        elif ws2.cell(i,2).value =='월요일':
            for j in range(1,9):
                ws2.cell(i,j).fill=c_fill2
                ws2.cell(i,j).font=font_bold
    #당일 환자명단 만들기
    date=datetime.date.today() #오늘날자 받아오기
    today=date.strftime("%Y-%m-%d") #오늘날자를 str type의 YYYY-mm-dd로 변경
    for i in range(2,ws.max_row+1):
        if ws.cell(i,1).value == today: #오늘 날자와 맞는 경우만 선택
            ws1.append([ws.cell(i,1).value,ws.cell(i,2).value,ws.cell(i,3).value,ws.cell(i,4).value,ws.cell(i,5).value,ws.cell(i,6).value,ws.cell(i,7).value])
    for i in range(2,ws1.max_row+1):
        for j in range(1,8):
            ws1.cell(i,j).border = border
            ws1.cell(i,j).alignment = align_center
    #당일환자명단 시간이 바뀌면 이중 밑줄 넣어주기
    for i in range(2,ws1.max_row+1):
        if ws1.cell(i,3).value != ws1.cell(i+1,3).value:
            for j in range(1,8):
                ws1.cell(i,j).border = border_btm_double
    #백신종류별환자수에 환자 많은날 표시
    for i in range(2,ws2.max_row+1):
        if ((ws2.cell(i,3).value >= sat_u_cnt) and (ws2.cell(i,2).value == '토요일')) or ((ws2.cell(i,3).value >= weekday_u_cnt) and (ws2.cell(i,2).value != '토요일')):
            for j in range(1,4):
                ws2.cell(i,j).border = border_thick
    wb.save(path_xlsx_2) #최종본.xlsx로 저장
    wb.close()
    root.geometry("220x50")
    root.title("변환완료")
    button = tkinter.Button(text = "OK", command = root.destroy)
    button.place(x=90, y=13)
    root.mainloop()
except:
    print("Unexpected error:", sys.exc_info()[0])
    root.geometry("200x50")
    root.title("!에러")
    button = tkinter.Button(text = "OK", command = root.destroy)
    button.place(x=90, y=13)  
    root.mainloop()